import pandas as pd
import json
import openpyxl
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from app.utils import db_utils, df_utils

def generate_turnusnokkel_for_turnus(turnus_name, turnus_set_id):
    """
    Generate turnusnøkkel Excel file for a specific turnus
    
    Args:
        turnus_name (str): Name of the turnus to generate key for
        turnus_set_id (int): ID of the turnus set
    
    Returns:
        dict: {'success': bool, 'filename': str, 'error': str}
    """
    try:
        # Get the turnus data for the specific turnus set
        df_manager = df_utils.DataframeManager(turnus_set_id)
        turnus_data = df_manager.turnus_data
        
        # Find the specific turnus data
        target_turnus_data = None
        for turnus_dict in turnus_data:
            if turnus_name in turnus_dict:
                target_turnus_data = turnus_dict[turnus_name]
                break
        
        if not target_turnus_data:
            return {'success': False, 'error': f'Turnus "{turnus_name}" not found in turnus set {turnus_set_id}'}
        
        # Get turnus set info for file paths
        turnus_set = db_utils.get_turnus_set_by_id(turnus_set_id)
        if not turnus_set:
            return {'success': False, 'error': f'Turnus set {turnus_set_id} not found'}
        
        year_identifier = turnus_set['year_identifier']
        
        # Define file paths based on turnus set
        base_path = f'app/static/turnusfiler/{year_identifier.lower()}'
        excel_template = f'{base_path}/turnusnøkkel_{year_identifier}_org.xlsx'
        json_file = f'{base_path}/turnuser_{year_identifier}.json'
        
        # Check if template exists
        if not os.path.exists(excel_template):
            return {'success': False, 'error': f'Template file not found: {excel_template}'}
        
        # Load the existing Excel file
        workbook = load_workbook(excel_template)
        sheet = workbook.active
        
        # Access the specific sheet
        sheet_name = 'Turnusnøkkel'
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            return {'success': False, 'error': f"Sheet '{sheet_name}' not found in the Excel file"}

        # Specify where to start inserting the data
        start_row = 51  # Row 51 for data insertion
        start_col = 1   # Column B (2nd column)

        # Process the turnus data and insert into Excel
        for uke_nr, ukedata in target_turnus_data.items():
            for dag_nr, dag_data in ukedata.items():
                try:
                    start_value = dag_data['tid'][0] if dag_data['tid'] and len(dag_data['tid']) > 0 else ''
                except (IndexError, KeyError, TypeError):
                    start_value = ''

                try:
                    end_value = dag_data['tid'][1] if dag_data['tid'] and len(dag_data['tid']) > 1 else ''
                except (IndexError, KeyError, TypeError):
                    end_value = ''
                
                # Create cell value
                if start_value and end_value:
                    cell_value = f'{start_value} - {end_value}'
                elif start_value:
                    cell_value = start_value
                else:
                    cell_value = ''

                # Set the value in the appropriate cell
                col_letter = get_column_letter(start_col + int(dag_nr))
                sheet[f"{col_letter}{start_row + int(uke_nr)}"] = cell_value
        
        # Generate filename and save
        filename = f"Turnusnøkkel_{turnus_name}_{year_identifier}.xlsx"
        output_path = f'{base_path}/generated/{filename}'
        
        # Create directory if it doesn't exist
        os.makedirs(f'{base_path}/generated', exist_ok=True)
        
        # Save the updated Excel file
        workbook.save(output_path)
        
        return {'success': True, 'filename': filename}
        
    except Exception as e:
        return {'success': False, 'error': f'Error generating turnusnøkkel: {str(e)}'}

# Legacy function for backward compatibility
def generate_all_turnusnokkel():
    """Generate turnusnøkkel for all turnus (legacy function)"""
    try:
        # Load the existing Excel file
        excel_file = 'turnusnøkkel_R25_org.xlsx'
        workbook = load_workbook(excel_file)
        sheet = workbook.active

        # Access the specific sheet
        sheet_name = 'Turnusnøkkel'
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            raise ValueError(f"Sheet '{sheet_name}' not found in the Excel file.")

        # Load the JSON file
        with open('turnuser_R25.json', 'r') as file:
            json_data = json.load(file)

        # Specify where to start inserting the data
        start_row = 51
        start_col = 1

        # Process all turnus data
        for turnus in json_data:
            for turnus_navn, turnus_data in turnus.items():
                # Load fresh workbook for each turnus
                workbook = load_workbook(excel_file)
                sheet = workbook[sheet_name]

                for uke_nr, ukedata in turnus_data.items():
                    for dag_nr, dag_data in ukedata.items():
                        try:
                            start_value = dag_data['tid'][0] if dag_data['tid'] and len(dag_data['tid']) > 0 else ''
                        except (IndexError, KeyError, TypeError):
                            start_value = ''

                        try:
                            end_value = dag_data['tid'][1] if dag_data['tid'] and len(dag_data['tid']) > 1 else ''
                        except (IndexError, KeyError, TypeError):
                            end_value = ''
                        
                        cell_value = f'{start_value} - {end_value}' if start_value and end_value else f"{start_value}"

                        col_letter_start = get_column_letter(start_col+int(dag_nr))
                        sheet[f"{col_letter_start}{start_row + int(uke_nr)}"] = cell_value
                
                # Save the updated Excel file
                filename = f"Turnusnøkkel_{turnus_navn}.xlsx"
                workbook.save(filename)

        print("Data successfully inserted into the specified location of the new Excel file.")
        return {'success': True}
        
    except Exception as e:
        print(f"Error: {str(e)}")
        return {'success': False, 'error': str(e)}