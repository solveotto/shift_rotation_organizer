import pandas as pd
import json
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Step 1: Load the existing Excel file
excel_file = 'turnusnøkkel_R25_org.xlsx'
workbook = load_workbook(excel_file)
sheet = workbook.active  # Select the active sheet, or specify the sheet name

# Step 2: Access a specific sheet by name
sheet_name = 'Turnusnøkkel'  # Replace with the actual sheet name
if sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
else:
    raise ValueError(f"Sheet '{sheet_name}' not found in the Excel file.")

# Step 2: Load the JSON file (assuming a 7x6 table structure)
with open('turnuser_R25.json', 'r') as file:
    json_data = json.load(file)

# Step 4: Specify where to start inserting the data
start_row = 51  # Row 2, for example
start_col = 1  # Column B (2nd column)


# Step 4: Extract 'start' and 'end' from each nested dictionary and insert them into cells
for turnus in json_data:
    for turnus_navn, turnus_data in turnus.items():
        # Step 1: Load the existing Excel file
        excel_file = 'turnusnøkkel_R25_org.xlsx'
        workbook = load_workbook(excel_file)
        sheet = workbook.active  # Select the active sheet, or specify the sheet name

        # Step 2: Access a specific sheet by name
        sheet_name = 'Turnusnøkkel'  # Replace with the actual sheet name
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            raise ValueError(f"Sheet '{sheet_name}' not found in the Excel file.")

        for uke_nr, ukedata in turnus_data.items():
            for dag_nr, dag_data in ukedata.items():
                
                    
                print(uke_nr, dag_nr, dag_data['tid'])

                try:
                    start_value = dag_data['tid'][0]
                except IndexError:
                    start_value = ''

                try:
                    end_value = dag_data['tid'][1]
                except IndexError:
                    end_value = ''
                
                
                cell_value = f'{start_value} - {end_value}' if start_value and end_value else f"{start_value}"

                # Set the 'start' value in the first column of this row
                col_letter_start = get_column_letter(start_col+int(dag_nr))
                sheet[f"{col_letter_start}{start_row + int(uke_nr)}"] = cell_value
            
        # Save the updated Excel file with a new name
        filename = f"Turnusnøkkel_{turnus_navn}.xlsx"
        workbook.save(filename)
                    


print("Data successfully inserted into the specified location of the new Excel file.")